import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.schemas.ata import Ata

class AtaExcelService:
    # Hardcoded questions to mirror frontend (since they are static in JS)
    AUDIT_DATA = [
        {
            "name": "Balanças Rodoviárias",
            "questions": [
                "01. As Balanças de Cana estão operando em condições Normais?",
                "02. Os Módulos da Balança de Cana estão ligados no No Break?",
                "03. A balança encontra-se limpa nos intervalos, cabeceira e células?",
                "04. Os resíduos estão sendo descartados corretamente nas lixeiras?",
                "05. Atestado de Autorização do INMETRO está atualizado e visível no local?",
                "06. Certificados de Calibração da Balança estão atualizados e visíveis?",
                "07. O turno de trabalho está completo?",
                "08. A Balança operou sem receber reclamação de outra área?",
                "09. O apontamento de Paradas foi realizado corretamente no PIMS?",
                "10. A troca de turno aconteceu no local de trabalho?",
                "11. O processo ocorreu sem a necessidade de Lançamento Manual de pesagem?",
                "12. Processo fluiu de forma contínua, sem paradas por falta de cana?",
                "13. Foi registrado algum problema (Quebra, Tara padrão, Erro Leitor)?"
            ]
        },
        {
            "name": "Laboratório PCTS",
            "questions": [
                "01. A coroa da Sonda Oblíqua está em bom estado de conservação (afiada e íntegra)?",
                "02. A Sonda Oblíqua opera com ausência de vazamento, vibração e ruído anormal?",
                "03. A data da última lubrificação dos equipamentos está no prazo?",
                "04. Foi realizada assepsia no Desintegrador, Homogeneizador e Esteiras?",
                "05. As facas do Desintegrador estão íntegras, afiadas e sem arredondamento?",
                "06. Os martelos do Desintegrador estão íntegros e sem arredondamento?",
                "07. O Desintegrador e Homogeneizador operam sem vibração e ruído anormal?",
                "08. A Prensa Hidráulica, Bomba de Óleo e Conexões estão sem vazamentos?",
                "09. Sonda, Desintegrador, Homogeneizador e Extrator encontram-se limpos?",
                "10. As balanças do setor de PCTS estão limpas e aferidas?",
                "11. Os demais equipamentos (Extrator, Centrífuga) funcionam normalmente?",
                "12. Pisos internos e externos limpos e livres de materiais inadequados?",
                "13. Bancadas e recipientes estão limpos e etiquetados nos devidos lugares?",
                "14. O refrigerador está limpo externamente e internamente?",
                "15. A temperatura interna do Laboratório está na faixa de 20º ± 3ºC?",
                "16. Equipamentos analíticos e TRDs estão funcionando sem problemas?",
                "17. Reagentes e soluções estão preparadas, identificadas e dentro da validade?",
                "18. A troca de turno aconteceu no local de trabalho?",
                "19. Foi evidenciado algum desvio adicional nas operações?",
                "20. O processo analítico fluiu sem paradas no laboratório?"
            ]
        },
        {
            "name": "Laboratório Industrial",
            "questions": [
                "01. Equipamentos Analíticos e outros estão em condições Normais de Operação?",
                "02. Os equipamentos estão limpos?",
                "03. Os reagentes e as soluções para as análises estão preparadas?",
                "04. A temperatura do Local de Trabalho se encontra em 20°C ± 3°C?",
                "05. As bancadas estão limpas e organizadas?",
                "06. O piso interno está limpo?",
                "07. O turno de Trabalho está completo?",
                "08. O Plano da Qualidade Analítico foi atendido em todas as áreas?",
                "09. O laboratório recebeu alguma reclamação da área do Processo?",
                "10. O Processo deixou de cumprir as Amostragens da Destilaria?",
                "11. A troca de turno aconteceu no local de trabalho?",
                "12. Processo parou ou sofreu interrupções no laboratório?",
                "13. Algum problema adicional foi registrado no Laboratório?",
                "14. As chaves dos amostradores estão disponíveis no local adequado?"
            ]
        }
    ]

    @staticmethod
    def generate_report(ata: Ata) -> io.BytesIO:
        wb = Workbook()

        # --- CAPA ---
        ws_capa = wb.active
        ws_capa.title = "Resumo da Ata"

        # Styles
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='005A7E', end_color='005A7E', fill_type='solid')
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        # Header
        ws_capa['A1'] = "ATA DIGITAL - DASHBOARD OPERACIONAL"
        ws_capa.merge_cells('A1:B1')
        ws_capa['A1'].font = title_font
        ws_capa['A1'].fill = header_fill
        ws_capa['A1'].alignment = center_align

        # Metadata
        # Use simple cell assignment to avoid merging issues if rows shift
        ws_capa['A3'] = "Data do Turno:"
        ws_capa['B3'] = str(ata.date)
        ws_capa['A4'] = "Turno:"
        ws_capa['B4'] = ata.shift
        ws_capa['A5'] = "Unidade:"
        ws_capa['B5'] = ata.unit
        ws_capa['A6'] = "Responsável:"
        ws_capa['B6'] = ata.responsible or "N/A"
        ws_capa['A7'] = "KPI Score:"
        ws_capa['B7'] = f"{ata.kpi_score}%"

        for row in range(3, 8):
            cell = ws_capa.cell(row=row, column=1)
            cell.font = bold_font

        ws_capa.column_dimensions['A'].width = 25
        ws_capa.column_dimensions['B'].width = 40

        # --- CHECKLIST TABS ---
        for cat_idx, category in enumerate(AtaExcelService.AUDIT_DATA):
            ws = wb.create_sheet(title=category["name"][:30]) # Excel limits sheet name length

            # Headers
            headers = ["ID", "Questão", "Resposta", "Tipo SBAR", "Situação (S)", "Background (B)", "Avaliação (A)", "Recomendação (R)"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = center_align

            # Data
            questions = category["questions"]
            cat_answers = ata.answers.get(cat_idx, {})
            cat_plans = ata.action_plans.get(cat_idx, {})

            for q_idx, q_text in enumerate(questions):
                row_idx = q_idx + 2

                # Answer
                # Pydantic dict keys might be int or str depending on serialization, cast to int safely
                ans = cat_answers.get(q_idx)

                # Plan
                plan = cat_plans.get(q_idx)

                # Write Row
                ws.cell(row=row_idx, column=1, value=q_idx + 1)
                ws.cell(row=row_idx, column=2, value=q_text)
                ws.cell(row=row_idx, column=3, value=ans if ans else "-")

                if plan:
                    ws.cell(row=row_idx, column=4, value=plan.type)
                    ws.cell(row=row_idx, column=5, value=plan.s)
                    ws.cell(row=row_idx, column=6, value=plan.b)
                    ws.cell(row=row_idx, column=7, value=plan.a)
                    ws.cell(row=row_idx, column=8, value=plan.r)

            # Adjust Widths
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 15
            for col in ['D', 'E', 'F', 'G', 'H']:
                ws.column_dimensions[col].width = 25

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
